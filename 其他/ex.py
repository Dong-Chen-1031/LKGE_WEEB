from openpyxl import workbook,load_workbook

import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore

cred = credentials.Certificate('E:\\lkge-web.json')


firebase_admin.initialize_app(cred)

# 初始化firestore
db = firestore.client()

wb = load_workbook("C:/Users/User/dasktop/林小資優通 (回覆).xlsx")

ws = wb.active



for i in range(2,4,1):
    print(str(i)+'：'+ws['B'+str(i)].value+' - '+ws['c'+str(i)].value+' - '+ws['D'+str(i)].value)
    path = "目錄/學生"
    doc_ref = db.document(path)
    doc = {ws['D'+str(i)].value:ws['C'+str(i)].value}
    doc_ref.update(doc)